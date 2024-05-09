import time
from datetime import datetime, timedelta
from threading import Thread

import openpyxl

from module import leidianDevice, image_recog, textOCR
from module.textOCR import TEXT_TYPE
from openpyxl import Workbook
from openpyxl import load_workbook

class Auto_record_Thread(Thread):
    def __init__(self, device_name):
        super().__init__()
        self.dev_name = device_name
        self.running = True
        self.excel_file = "gold_record.xlsx"

    def stop_thread(self):
        self.running = False

    def run(self):
        """
        自动选择副本： 站街状态 -> 开始秘境
        """
        # 1.检测当前（站街）是否能找到背包
        while self.running:
            # path = leidianDevice.get_current_pic_path(self.dev_name)
            # (img, x, y) = image_recog.get_target_point(path, '背包.pic.png', 0.8)
            # if img is not None:
            #     print("找到背包按钮")
            #     leidianDevice.tap(self.dev_name, x, y)
            #     time.sleep(2)
            #     break
            # else:
            #     print("未找到背包按钮")
            # time.sleep(1)
            break

        path = leidianDevice.get_current_pic_path(self.dev_name)
        # time.sleep(1)
        res = textOCR.get_num(path, TEXT_TYPE.GOLD)

        # 写入excel
        self.write_value_excel(name="name2", value=2)

    def write_value_excel(self, name, value):  # openpyxl库储存数据到excel
        # 创建一个工作簿
        try:
            wb = load_workbook(self.excel_file)
            print("已打开现有的 Excel 文件。")
        except FileNotFoundError:
            # 如果文件不存在，则创建一个新的 Excel 文件
            wb = Workbook()
            print("创建了一个新的 Excel 文件。")
            wb.save(self.excel_file)

            ws = wb.active

            ws.cell(row=1, column=1, value="")
            wb.save(self.excel_file)
            self.add_10_date_in_row_1(wb=wb, from_index=2)
            self.set_all_columns_width(wb=wb)

        # 激活默认的工作表
        ws = wb.active

        row_index = None
        for row in ws.iter_cols(values_only=True):
            if name in row:
                row_index = row.index(name)
                break
        if row_index:
            print(f"找到条件值 \"{name}\" 所在的列：{row_index}")
        else:
            print(f"未找到条件值 \"{name}\"")
            max_row = ws.max_row
            row_index = max_row+1
            ws.cell(row=row_index, column=1, value=name)
             # 保存工作簿
            wb.save(self.excel_file)


        date = str(datetime.now().date())
        col_index = None
        for cell in ws[1]:
            # 如果单元格的值是 'date'
            value_str = str(cell.value).split(" ")[0]
            if value_str == date:
                col_index = cell.column
                print(f"{date} 在第一行中的列索引为: {col_index}")
                break
        else:
            print("未在第一行中找到 'date'")
            last_column = ws.max_column+1
            col_index = last_column
            self.add_10_date_in_row_1(wb=wb, from_index=last_column)

        # 最后一列
        # last_column = ws.max_column
        # last_column_label = openpyxl.utils.get_column_letter(last_column)

        # 写入数据到单元格
        next_cell = ws.cell(row=row_index, column=col_index, value=value)

        # 保存工作簿
        wb.save(self.excel_file)

    def add_10_date_in_row_1(self, wb, from_index):
        ws = wb.active
        current_date = datetime.now().date()
        for i in range(10):
            date_to_add = current_date + timedelta(days=i)
            ws.cell(row=1, column=i + from_index, value=date_to_add)
        wb.save(self.excel_file)

    def set_all_columns_width(self, wb, width=12):
        ws = wb.active
        for column in ws.columns:
            ws.column_dimensions[column[0].column_letter].width = width
        # 保存工作簿
        wb.save(self.excel_file)


if __name__ == '__main__':
    c = Auto_record_Thread('雷电模拟器')
    c.start()
